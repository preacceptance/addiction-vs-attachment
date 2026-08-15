#!/usr/bin/env python3
"""
Shape the drawn IRR samples into the workbooks the human coders use.

Tab layout: Consensus | one tab per coder | Legend.
- Coder tabs: identity columns + text + that coder's code and justification.
  The code column has a dropdown sourced from Legend!$A$1:$A$4.
- Consensus tab: pulls both coders' entries by formula (live once they type),
  computes agreement, and leaves the notes and final_code columns blank for
  reconciliation. final_code also gets the Legend dropdown.
- Legend tab: the four codes, one per row, the dropdowns' source range.

Reads the blind draws from draw_irr_24.py and REWRITES the same filenames.

Usage:  python3 build_irr_sheets.py
"""
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent / ".." / "modified_data"
CODES = ["Addiction", "Attachment", "Both", "Neither"]

def build(src: str, id_cols: list[str], text_col: str) -> None:
    df = pd.read_excel(HERE / src)
    n = len(df)
    wb = Workbook()
    wb.remove(wb.active)

    def dv_for(ws):
        dv = DataValidation(type="list", formula1="=Legend!$A$1:$A$4", allow_blank=True)
        ws.add_data_validation(dv)
        return dv

    def write_header(ws, cols):
        for j, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=j, value=c)
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"

    def set_widths(ws, cols):
        for j, c in enumerate(cols, 1):
            w = 60 if c == text_col else 22 if "justification" in c or "notes" in c else 14
            ws.column_dimensions[get_column_letter(j)].width = w

    base_cols = ["row"] + id_cols + [text_col]

    def coder_tab(name, code_col, just_col):
        cols = base_cols + [code_col, just_col]
        ws = wb.create_sheet(name)
        write_header(ws, cols)
        set_widths(ws, cols)
        dv = dv_for(ws)
        code_letter = get_column_letter(len(base_cols) + 1)
        for i, r in df.iterrows():
            row = i + 2
            for j, c in enumerate(base_cols, 1):
                v = r[c]
                cell = ws.cell(row=row, column=j, value=None if pd.isna(v) else v)
                if c == text_col:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        dv.add(f"{code_letter}2:{code_letter}{n+1}")
        return code_letter

    # Consensus is created first so it is the tab the workbook opens on
    cons_cols = base_cols + ["code_rater1", "justification_rater1", "code_rater2",
                             "justification_rater2", "agreement",
                             "notes_rater1", "notes_rater2", "final_code"]
    cons = wb.create_sheet("Consensus")
    write_header(cons, cons_cols)
    set_widths(cons, cons_cols)

    it_letter = coder_tab("Rater1", "code_rater1", "justification_rater1")
    om_letter = coder_tab("Rater2", "code_rater2", "justification_rater2")

    k = len(base_cols)
    L = {name: get_column_letter(k + off + 1) for off, name in enumerate(
        ["code_rater1", "justification_rater1", "code_rater2", "justification_rater2",
         "agreement", "notes_rater1", "notes_rater2", "final_code"])}
    dv_cons = dv_for(cons)
    for i, r in df.iterrows():
        row = i + 2
        for j, c in enumerate(base_cols, 1):
            v = r[c]
            cell = cons.cell(row=row, column=j, value=None if pd.isna(v) else v)
            if c == text_col:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        jl = get_column_letter(k + 2)  # coder-tab justification column
        cons[f"{L['code_rater1']}{row}"] = f"=Rater1!{it_letter}{row}"
        cons[f"{L['justification_rater1']}{row}"] = f"=Rater1!{jl}{row}"
        cons[f"{L['code_rater2']}{row}"] = f"=Rater2!{om_letter}{row}"
        cons[f"{L['justification_rater2']}{row}"] = f"=Rater2!{jl}{row}"
        a, b = L["code_rater1"], L["code_rater2"]
        cons[f"{L['agreement']}{row}"] = (
            f'=IF(OR({a}{row}="",{b}{row}=""),"",IF({a}{row}={b}{row},"agree","DISAGREE"))')
    dv_cons.add(f"{L['final_code']}2:{L['final_code']}{n+1}")

    lg = wb.create_sheet("Legend")
    for i, c in enumerate(CODES, 1):
        lg.cell(row=i, column=1, value=c)
    lg.cell(row=1, column=3, value="Codes for the dropdowns. Code on your own tab only; "
                                   "reconcile disagreements in Consensus (final_code) without "
                                   "discussing who coded what first.")

    wb.save(HERE / src)
    print(f"rebuilt {src}: tabs {[s.title for s in wb.worksheets]}, {n} rows")

if __name__ == "__main__":
    build("Legal IRR 24cases STRATIFIED N150.xlsx", ["unit_id", "case", "pdf_page"], "text")
    build("Media IRR 24cases STRATIFIED N150.xlsx", ["unit_id", "pdf", "page", "headline"], "para_text")
